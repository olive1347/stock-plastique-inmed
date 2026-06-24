"""
GestStock INMED — Gestion du Stock Plastique & Consommables
Lecture intelligente des fichiers de demande Excel avec propagation des désignations vides
Édition directe en mode tableur (type Excel) stable et persistant
Restriction de sécurité par adresse IP (Réseau local INMED / VPN)
"""

#import json
import re
import io
import unicodedata
from pathlib import Path
import pandas as pd
import openpyxl
import requests
import streamlit as st
from datetime import datetime

# Configuration globale de la fenêtre de navigation (impérativement en premier)
st.set_page_config(page_title="GestStock INMED", page_icon="🧪", layout="wide")

# ══════════════════════════════════════════════════════════════
# CONFIGURATION ET CONSTANTES (SÉCURITÉ ET IP)
# ══════════════════════════════════════════════════════════════
#STATE_FILE = "stock_state.json"
EXCEL_FILE = "Fiche demande labo plastique 2025-2026.xlsx"

try:
    groq_key = st.secrets.get("GROQ_API_KEY", st.secrets.get("groq_api_key", ""))
except Exception:
    groq_key = ""

# CONFIGURATION SÉCURISÉE DE L'ACCÈS RÉSEAU
CONFIG = {
    "groq_api_key": groq_key,
    "groq_url"    : "https://api.groq.com/openai/v1/chat/completions",
    "model"       : "llama-3.1-8b-instant",
    "admin_pass"  : "inmed2026", # Code d'accès pour l'espace d'Olivier
    
    # Activer/Désactiver la restriction IP globale (True = Actif, False = Tout le monde accède)
    "ip_restriction_enabled": True,  
    
    # Liste blanche des préfixes d'IP autorisés (Réseau local & VPN)
    "allowed_ip_prefixes": [
        "127.0.0.1",    # Test local IPv4
        "::1",          # Test local IPv6
        "localhost",    # Domaine local
        "10.",          # Réseau privé classique (très fréquent en institut/Inserm)
        "172.16.",      # Plage privée Classe B (172.16.x.x à 172.31.x.x)
        "172.17.",
        "172.18.",
        "172.19.",
        "172.20.",
        "172.21.",
        "172.22.",
        "172.23.",
        "172.24.",
        "172.25.",
        "172.26.",
        "172.27.",
        "172.28.",
        "172.29.",
        "172.30.",
        "172.31.",
        "192.168.",     # Réseau privé classique Classe C
        "139.124.",     # Exemple d'adresse académique AMU
        "147.210."      # Exemple d'adresse académique Inserm
    ]
}

OFFICIAL_CATEGORIES = [
    "FILTRATION",
    "HISTO ET ELECTROPHY",
    "TUBES",
    "CULTURE CELLULAIRE",
    "POINTES OU TIPS",
    "SERINGUES ET AIGUILLES",
    "BACTÉRIO",
    "PCR",
    "DIVERS",
    "PESÉES"
]

# ══════════════════════════════════════════════════════════════
# FONCTIONS DE SÉCURITÉ RÉSEAU (DÉTECTION IP)
# ══════════════════════════════════════════════════════════════
#def get_client_ip():
    """Récupère l'adresse IP réelle du chercheur à partir des en-têtes HTTP de Streamlit"""
    try:
        headers = st.context.headers
        if "X-Forwarded-For" in headers:
            # Récupère le premier élément de la chaîne (l'IP d'origine du client)
            return headers["X-Forwarded-For"].split(",")[0].strip()
        if "X-Real-IP" in headers:
            return headers["X-Real-IP"].strip()
    except Exception:
        pass
        
    # Méthode alternative de repli
    try:
        from streamlit.web.server.websocket_headers import _get_websocket_headers
        headers = _get_websocket_headers()
        if headers:
            if "X-Forwarded-For" in headers:
                return headers["X-Forwarded-For"].split(",")[0].strip()
            if "X-Real-IP" in headers:
                return headers["X-Real-IP"].strip()
    except Exception:
        pass
        
    return "127.0.0.1" # Valeur par défaut pour le développement local

def is_ip_authorized(client_ip, allowed_prefixes):
    """Vérifie si l'adresse IP commence par l'un des préfixes réseau autorisés"""
    if not client_ip:
        return False
    client_ip = client_ip.strip()
    return any(client_ip.startswith(prefix) for prefix in allowed_prefixes)

# ══════════════════════════════════════════════════════════════
# FONCTIONS DE NETTOYAGE ET D'ANALYSE SEMANTIQUE
# ══════════════════════════════════════════════════════════════
#def clean_text(text):
    if not text:
        return ""
    s = str(text).lower()
    s = "".join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return re.sub(r'[^a-z]', '', s)

def detect_category(designation_val, info_val, cdt_val, ref_val):
    if not designation_val:
        return None
        
    des_clean = designation_val.strip().upper()
    if info_val or cdt_val or ref_val:
        return None
        
    des_clean = re.sub(r'^\d+[\s\.\-)]+', '', des_clean)
    
    cats_map = {
        "FILTRATION": ["FILTRATION", "FILTER"],
        "HISTO ET ELECTROPHY": ["HISTO", "ELECTROPHY", "ELECTROPHYS"],
        "TUBES": ["TUBES", "TUBE"],
        "CULTURE CELLULAIRE": ["CULTURE", "CELLULAIRE"],
        "POINTES OU TIPS": ["POINTES", "TIPS", "POINTE", "TIP"],
        "SERINGUES ET AIGUILLES": ["SERINGUES", "AIGUILLES", "SERINGUE", "AIGUILLE"],
        "BACTÉRIO": ["BACTERIO", "BACTÉRIO"],
        "PCR": ["PCR"],
        "DIVERS": ["DIVERS"],
        "PESÉES": ["PESEE", "PESEES", "PESÉES", "PESÉE"]
    }
    
    for official_name, keywords in cats_map.items():
        if any(kw in clean_text(des_clean) for kw in keywords):
            return official_name
            
    return None

def get_cell_value(ws, row, col):
    cell = ws.cell(row=row, column=col)
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return cell.value

# ══════════════════════════════════════════════════════════════
# PARSER EXCEL INTELLIGENT
# ══════════════════════════════════════════════════════════════
#def import_excel_catalog(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        catalog = []
        current_category = "DIVERS"
        last_valid_designation = ""
        
        for r in range(5, ws.max_row + 1):
            raw_des = get_cell_value(ws, r, 1)
            raw_info = get_cell_value(ws, r, 2)
            raw_cdt = get_cell_value(ws, r, 3)
            raw_ref = get_cell_value(ws, r, 5)
            
            if not raw_des and not raw_info and not raw_cdt and not raw_ref:
                continue
                
            cat_detected = detect_category(raw_des, raw_info, raw_cdt, raw_ref)
            if cat_detected:
                current_category = cat_detected
                last_valid_designation = ""
                continue
                
            if raw_des and str(raw_des).strip() != "":
                last_valid_designation = str(raw_des).strip()
                
            final_designation = last_valid_designation
            if raw_info and str(raw_info).strip() != "":
                if final_designation:
                    final_designation += f" - {str(raw_info).strip()}"
                else:
                    final_designation = str(raw_info).strip()
            
            if not final_designation:
                continue
                
            catalog.append({
                "categorie": current_category,
                "designation": final_designation,
                "cdt": str(raw_cdt).strip() if raw_cdt else "Unité",
                "ref_fab": str(raw_ref).strip() if raw_ref else "N/A",
                "stock": 50,
                "seuil_alerte": 10
            })
            
        return {"success": True, "data": catalog}
    except Exception as e:
        return {"success": False, "error": str(e)}

# ══════════════════════════════════════════════════════════════
# GESTION DES DONNÉES EN SESSION ET SAUVEGARDE
# ══════════════════════════════════════════════════════════════
#def save_state_to_file():
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(st.session_state.db_stock, f, ensure_ascii=False, indent=4)

def load_state():
    if Path(STATE_FILE).exists():
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def apply_editor_changes(df_original, state_changes):
    res_df = df_original.copy()
    
    for row_idx_str, edits in state_changes.get("edited_rows", {}).items():
        row_idx = int(row_idx_str)
        if row_idx < len(res_df):
            for col, val in edits.items():
                res_df.at[row_idx, col] = val
                
    for addition in state_changes.get("added_rows", []):
        new_row = {col: addition.get(col, "") for col in res_df.columns}
        if not new_row.get("stock"): 
            new_row["stock"] = 0
        if not new_row.get("seuil_alerte"): 
            new_row["seuil_alerte"] = 10
        if not new_row.get("categorie"):
            new_row["categorie"] = "DIVERS"
        res_df = pd.concat([res_df, pd.DataFrame([new_row])], ignore_index=True)
        
    deletions = sorted([int(idx) for idx in state_changes.get("deleted_rows", [])], reverse=True)
    for idx in deletions:
        if idx < len(res_df):
            res_df = res_df.drop(res_df.index[idx]).reset_index(drop=True)
            
    return res_df

# Initialisation des états de session
if "db_stock" not in st.session_state:
    st.session_state.db_stock = load_state()

if "commandes_attente" not in st.session_state:
    st.session_state.commandes_attente = []

if "editor_version" not in st.session_state:
    st.session_state.editor_version = 0

# ══════════════════════════════════════════════════════════════
# CONTRÔLE DE SÉCURITÉ PAR ADRESSE IP
# ══════════════════════════════════════════════════════════════
#client_ip = get_client_ip()
is_authorized = is_ip_authorized(client_ip, CONFIG["allowed_ip_prefixes"])

# BLOCAGE SI HORS-RÉSEAU
if CONFIG["ip_restriction_enabled"] and not is_authorized:
    st.error("⛔ Accès sécurisé restreint")
    st.markdown(
        f"""
        ### 🔒 Cette application est exclusivement réservée au personnel de l'INMED.
        
        Pour pouvoir accéder au catalogue, passer vos commandes ou gérer l'inventaire :
        1. **Si vous êtes à l'institut :** Assurez-vous d'être connecté au réseau filaire (Ethernet) ou au réseau Wi-Fi officiel de l'institut.
        2. **Si vous êtes à distance (Télétravail, déplacement) :** Vous devez impérativement démarrer et vous connecter au **VPN officiel** (Inserm ou Aix-Marseille Université) avant d'actualiser cette page.
        
        ---
        📍 **Votre adresse IP de connexion actuelle :** `{client_ip}` (Non autorisée)
        """
    )
    st.stop() # Arrêt immédiat pour les accès externes non autorisés

# ══════════════════════════════════════════════════════════════
# INTERFACE GRAPHIQUE STANDARD (UNIQUEMENT SI AUTORISÉ)
# ══════════════════════════════════════════════════════════════
#st.title("🧪 GestStock INMED — Gestionnaire de Stock")

# Menu de gauche
onglet = st.sidebar.radio("🧭 Menu principal", ["👋 Espace Chercheurs (Demandes)", "🔑 Espace Administration (Olivier)"])

# ──────────────────────────────────────────────────────────────
# ONGLET 1 : ESPACE CHERCHEURS (DEMANDES & CATALOGUE)
# ──────────────────────────────────────────────────────────────
#if onglet == "👋 Espace Chercheurs (Demandes)":
    st.header("👋 Passer une commande de consommables plastiques")
    
    if not st.session_state.db_stock:
        st.warning("⚠️ L'inventaire est actuellement vide. Olivier doit d'abord synchroniser le fichier Excel dans son espace.")
    else:
        tab_demande, tab_cat = st.tabs(["📝 Formulaire de Demande Rapide", "📋 Consulter l'inventaire"])
        
        with tab_demande:
            st.write("Remplissez vos besoins ci-dessous pour soumettre votre demande à Olivier.")
            
            # Informations chercheur
            col_u1, col_u2 = st.columns(2)
            with col_u1:
                nom_demande = st.text_input("Nom & Prénom :", placeholder="ex: Sophie Durand")
            with col_u2:
                equipe_demande = st.text_input("Destination / Équipe :", placeholder="ex: Équipe Neuro-Physio")
                
            # Sélection de l'article
            st.write("---")
            st.write("**Ajouter des articles à votre demande :**")
            
            categories_existantes = sorted(list(set(item["categorie"] for item in st.session_state.db_stock)))
            cat_choisie = st.selectbox("Sélectionnez une catégorie :", categories_existantes)
            
            articles_filtrés = [item for item in st.session_state.db_stock if item["categorie"] == cat_choisie]
            options_articles = {item["designation"]: item for item in articles_filtrés}
            
            art_choisi_nom = st.selectbox("Sélectionnez l'article :", list(options_articles.keys()))
            quantite_demandee = st.number_input("Quantité souhaitée :", min_value=1, value=1, step=1)
            
            if "panier" not in st.session_state:
                st.session_state.panier = []
                
            if st.button("➕ Ajouter au panier"):
                art_details = options_articles[art_choisi_nom]
                st.session_state.panier.append({
                    "designation": art_details["designation"],
                    "categorie": art_details["categorie"],
                    "cdt": art_details["cdt"],
                    "ref_fab": art_details["ref_fab"],
                    "quantite": int(quantite_demandee)
                })
                st.success(f"Ajouté : {quantite_demandee}x {art_choisi_nom}")
                st.rerun()
                
            # Affichage et validation du panier
            if st.session_state.panier:
                st.write("---")
                st.subheader("🛒 Votre Panier de Demande")
                df_panier = pd.DataFrame(st.session_state.panier)
                st.dataframe(df_panier, use_container_width=True)
                
                col_pan1, col_pan2 = st.columns(2)
                with col_pan1:
                    if st.button("🚀 Soumettre la demande à Olivier", type="primary", use_container_width=True):
                        if nom_demande and equipe_demande:
                            st.session_state.commandes_attente.append({
                                "id": len(st.session_state.commandes_attente) + 1,
                                "chercheur": nom_demande,
                                "equipe": equipe_demande,
                                "date": datetime.today().strftime("%d/%m/%Y à %H:%M"),
                                "items": st.session_state.panier.copy(),
                                "statut": "En attente"
                            })
                            st.session_state.panier = []
                            st.balloons()
                            st.success("✅ Votre demande a bien été transmise à Olivier ! Vous pouvez aller récupérer vos consommables.")
                            st.rerun()
                        else:
                            st.error("⚠️ Veuillez renseigner votre Nom et votre Équipe avant de valider.")
                with col_pan2:
                    if st.button("🗑️ Vider le panier", use_container_width=True):
                        st.session_state.panier = []
                        st.rerun()
                        
        with tab_cat:
            st.subheader("📋 État général de l'inventaire")
            df_public = pd.DataFrame(st.session_state.db_stock)
            df_public_display = df_public[["categorie", "designation", "cdt", "ref_fab", "stock"]].copy()
            df_public_display.columns = ["Catégorie", "Désignation", "Conditionnement", "Réf Fabricant", "En Stock"]
            st.dataframe(df_public_display, use_container_width=True)

# ──────────────────────────────────────────────────────────────
# ONGLET 2 : ESPACE D'ADMINISTRATION (OLIVIER)
# ──────────────────────────────────────────────────────────────
#elif onglet == "🔑 Espace Administration (Olivier)":
    st.header("🔑 Espace Gestionnaire de Stock")
    
    password = st.text_input("Code de sécurité administrateur :", type="password")
    
    if password == CONFIG["admin_pass"]:
        st.success("🔓 Accès administrateur autorisé.")
        
        adm_tab1, adm_tab2, adm_tab3, adm_tab4 = st.tabs([
            "📥 File d'attente des commandes", 
            "✏️ Éditeur de Stock Interactif (Type Excel)", 
            "🔄 Synchronisation & Réinitialisation",
            "🛡️ Statut Sécurité Réseau & IP"
        ])
        
        # ─── SOUS-ONGLET 1 : LES DEMANDES EN ATTENTE ───
        with adm_tab1:
            st.subheader("📦 Commandes à distribuer")
            attente = [c for c in st.session_state.commandes_attente if c["statut"] == "En attente"]
            
            if not attente:
                st.info("☕ Aucune commande en attente pour le moment.")
            else:
                for cmd in attente:
                    with st.expander(f"📋 Commande #{cmd['id']} - {cmd['chercheur']} ({cmd['equipe']}) - {cmd['date']}"):
                        df_items_cmd = pd.DataFrame(cmd["items"])
                        st.dataframe(df_items_cmd[["designation", "cdt", "ref_fab", "quantite"]], use_container_width=True)
                        
                        col_cmd1, col_cmd2 = st.columns(2)
                        with col_cmd1:
                            if st.button("✅ Valider & déduire du stock physique", key=f"val_{cmd['id']}", use_container_width=True):
                                error_flag = False
                                for req_item in cmd["items"]:
                                    match = next((item for item in st.session_state.db_stock if item["designation"] == req_item["designation"]), None)
                                    if match:
                                        if match["stock"] < req_item["quantite"]:
                                            st.error(f"Stock insuffisant pour {req_item['designation']} (Requis: {req_item['quantite']}, En stock: {match['stock']})")
                                            error_flag = True
                                
                                if not error_flag:
                                    for req_item in cmd["items"]:
                                        match = next((item for item in st.session_state.db_stock if item["designation"] == req_item["designation"]), None)
                                        if match:
                                            match["stock"] -= req_item["quantite"]
                                    cmd["statut"] = "Délivrée"
                                    save_state_to_file()
                                    st.success("✅ Stock déduit et commande archivée !")
                                    st.rerun()
                        with col_cmd2:
                            if st.button("❌ Rejeter / Annuler", key=f"rej_{cmd['id']}", use_container_width=True):
                                cmd["statut"] = "Annulée"
                                st.warning("Commande annulée.")
                                st.rerun()

        # ─── SOUS-ONGLET 2 : EDITEUR DIRECT DU STOCK (TYPE EXCEL) ───
        #        with adm_tab2:
            st.subheader("✏️ Édition instantanée de vos données")
            st.write("Double-cliquez sur n'importe quelle case ci-dessous pour modifier directement le stock, le nom, la catégorie ou la référence.")
            
            df_editor = pd.DataFrame(st.session_state.db_stock)
            if df_editor.empty:
                df_editor = pd.DataFrame(columns=["categorie", "designation", "cdt", "ref_fab", "stock", "seuil_alerte"])
                
            editor_key = f"stock_editor_v{st.session_state.editor_version}"
            
            edited_df = st.data_editor(
                df_editor,
                column_config={
                    "categorie": st.column_config.SelectboxColumn("Catégorie", options=OFFICIAL_CATEGORIES, width="medium"),
                    "designation": st.column_config.TextColumn("Désignation Produit", required=True, width="large"),
                    "cdt": st.column_config.TextColumn("Conditionnement", width="medium"),
                    "ref_fab": st.column_config.TextColumn("Réf Fabricant", width="medium"),
                    "stock": st.column_config.NumberColumn("Quantité en Stock", min_value=0, required=True),
                    "seuil_alerte": st.column_config.NumberColumn("Seuil Alerte", min_value=0, required=True),
                },
                num_rows="dynamic",
                use_container_width=True,
                key=editor_key
            )
            
            if st.button("💾 Enregistrer toutes les modifications du tableau", type="primary", use_container_width=True):
                if editor_key in st.session_state:
                    state_changes = st.session_state[editor_key]
                    final_df = apply_editor_changes(df_editor, state_changes)
                else:
                    final_df = edited_df
                
                st.session_state.db_stock = final_df.to_dict(orient="records")
                save_state_to_file()
                st.session_state.editor_version += 1
                st.success("✅ Base de données mise à jour et sauvegardée avec succès !")
                st.rerun()

        # ─── SOUS-ONGLET 3 : PARSING EXCEL & REINITIALISATION ───
        #        with adm_tab3:
            st.subheader("🔄 Synchronisation avec le fichier Excel INMED")
            st.write(f"Cette fonction va lire le fichier **`{EXCEL_FILE}`** à la racine de votre projet GitHub.")
            
            if st.button("⚙️ Lancer la lecture automatique de l'Excel", use_container_width=True):
                if not Path(EXCEL_FILE).exists():
                    st.error(f"❌ Fichier `{EXCEL_FILE}` introuvable à la racine de votre répertoire. Assurez-vous qu'il est bien présent sur GitHub.")
                else:
                    with st.status("Lecture et analyse fine de votre Excel en cours...", expanded=True) as status:
                        res = import_excel_catalog(EXCEL_FILE)
                        if res["success"]:
                            st.session_state.db_stock = res["data"]
                            save_state_to_file()
                            st.session_state.editor_version += 1
                            status.update(label="✅ Fichier lu et analysé à 100% sans erreur !", state="complete")
                            st.success(f"Bravo ! {len(res['data'])} articles ont été correctement extraits.")
                            st.rerun()
                        else:
                            status.update(label="❌ Erreur durant la lecture", state="error")
                            st.error(res["error"])
                            
            st.write("---")
            st.subheader("⚠️ Réinitialisation complète")
            if st.button("🗑️ Effacer tout l'inventaire en cours", type="secondary", use_container_width=True):
                st.session_state.db_stock = []
                save_state_to_file()
                st.session_state.editor_version += 1
                st.warning("Tout l'inventaire a été remis à zéro.")
                st.rerun()

        # ─── SOUS-ONGLET 4 : SECURITE RESEAU & IP (SÉCURITÉ INMED) ───
        #        with adm_tab4:
            st.subheader("🛡️ Paramètres et journal de restriction réseau")
            st.write("Cet espace vous sert à gérer les règles de filtrage de l'application.")
            
            col_ip1, col_ip2 = st.columns(2)
            with col_ip1:
                st.metric("Votre adresse IP externe détectée", client_ip)
                st.metric("Résultat du filtrage", "✅ Connecté localement (Autorisé)" if is_authorized else "❌ Connexion extérieure (Bloqué)")
            with col_ip2:
                st.write("**Préfixes autorisés configurés :**")
                st.code(json.dumps(CONFIG["allowed_ip_prefixes"], indent=2), language="json")
            
            st.write("---")
            st.markdown(
                """
                💡 **Mémo administrateur :**
                * Les adresses IP locales de type `10.x.x.x` (qui équipent la majorité des bureaux et paillasses de l'INMED en filaire) sont automatiquement autorisées.
                * Les tunnels VPN académiques ou de l'Inserm sont également pris en charge via les plages d'IP académiques configurées.
                * Si un chercheur de confiance vous signale qu'il est bloqué alors qu'il est au laboratoire, demandez-lui de vous donner l'IP qui s'affiche sur son écran de blocage rouge, et rajoutez son préfixe dans l'objet `allowed_ip_prefixes` de la configuration du script Python.
                """
            )
                
    elif password != "":
        st.error("🔑 Code d'accès administrateur incorrect.")
